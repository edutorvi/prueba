# -*- coding: utf-8 -*-
from io import BytesIO#PARA REPORTLAB

from django.views.generic import TemplateView, CreateView, UpdateView, DeleteView, ListView
from apps.fotos.models import Foto, Tipofoto, Procedencia
from apps.fotos.forms import FotoForm,TipoFotoForm,ProcedenciaForm
from django.views import  View
from django.urls import reverse_lazy
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.http import JsonResponse
from django.core import serializers

from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
from reportlab.lib.styles import getSampleStyleSheet#estilosparrafo
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
#from reportlab.lib.pagesizes import LETTER
from reportlab.platypus import Table
from reportlab.pdfgen import canvas
#Librerias reportlab a usar:
from reportlab.platypus import (BaseDocTemplate, Frame,  
					NextPageTemplate, PageBreak, PageTemplate)
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4

from reportlab.lib.styles import ParagraphStyle

from reportlab.lib.units import mm
from reportlab.platypus import Image
from reportlab.platypus.flowables import Spacer


from apps.fotos.forms import UploadImageForm
from django.shortcuts import render
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
import os
from django.http import HttpResponseRedirect

from apps.fotos.pdf import Pdf, FooterCanvas


class Login5(TemplateView):
	template_name = "index.html"

def upload_image_view(request):
	#return HttpResponse('Hello, World!')
	if request.method == 'POST':
		form = UploadImageForm(request.POST, request.FILES)
		if form.is_valid():
			form.save()
			message = "Image uploaded succesfully!"
	else:
		form = UploadImageForm()

	#return render_to_response('fotos/formulariosubir.html', locals(), context_instance=RequestContext(request))
	#return render(request, 'fotos/formulariosubir.html', {'form': form})
	return render(request, 'fotos/formulariosubir.html', locals())
	#return render_to_response('add_company.html', {'form': form}, context)


class Home(TemplateView):
	template_name = "inicio.html"

class FooterCanvas(canvas.Canvas):

	def __init__(self, *args, **kwargs):
		canvas.Canvas.__init__(self, *args, **kwargs)
		self.pages = []

	def showPage(self):
		self.pages.append(dict(self.__dict__))
		self._startPage()

	def save(self):
		page_count = len(self.pages)
		for page in self.pages:
			self.__dict__.update(page)
			self.draw_canvas(page_count)
			canvas.Canvas.showPage(self)
		canvas.Canvas.save(self)

	def draw_canvas(self, page_count):
		page1 = "Pág %s de %s" % (self._pageNumber, page_count)
		page2 = "Gobierno Regional de Cajamarca"
		x1 = 128
		x2 = 500
		x2 = 461
		self.saveState()
		self.setStrokeColorRGB(0, 0, 0)
		self.setLineWidth(0.5)
		self.line(66, 71, A4[0] - 66, 71)#linea footer
		self.setFont('Times-Roman', 10)
		self.drawString(A4[0]-x1, 59, page1)#texto footer
		
		self.setFont('Times-Roman', 12)
		self.drawString(A4[0]-x2, 785, page2)#texto cabecera
		self.line(66, 780, A4[0] - 66, 780)#linea cabecera

		#self = canvas.Canvas(filename, bottomup=0)
		f=Image('static/img/logo_Region_Cajamarca2019.png', width=147, height=50)#imagen
		#f=Image('static/img/gr2019.png', width=73, height=25)#imagen
		f.drawOn(self, 75, 780)#posicion DE LA IMAGEN en el encabezado


		self.restoreState()
	
class Listar(TemplateView):
	template_name = "listar.html"
	def pdf_Platypus(request):
		response = HttpResponse(content_type='application/pdf')#tipo de rpta
		#pdf_name = "fotos1.pdf"  # nombre para la descarga
		#response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name #descarga con el nombre indicado
		buff = BytesIO() #almacena el doc
		#c = canvas.Canvas(buff)
		doc = SimpleDocTemplate(buff,
								pagesize=A4,
								rightMargin=50,#miderecha
								leftMargin=50,#mi izquierda
								topMargin=60,
								bottomMargin=70,#inferior hoja	                            
								)#plantilla del doc enlazado al buffer
		fotos = []#lista para generar doc
		styles = getSampleStyleSheet()#estilos
		style = styles['Title']
		style1= styles['Normal']
		style.spaceAfter = 0.3 * inch
		style.fontSize = 21
		style.fontName = "Times-Roman"
		style.textColor = "Blue"

		#style.alignment=TA_JUSTIFY
		"""p = ParagraphStyle('parrafos', 
						   
						   fontSize = 28,
						   fontName="Times-Roman")"""
		header = Paragraph("Listado de Fotos", style)#encabezado doc
		#header = Paragraph("Listado de Fotos", styles['Heading1'])#encabezado doc
		#header = Paragraph("Listado de Fotos", p)#encabezado doc
		
		"""f=Image('static/img/admin.png', width=30, height=30)
		f.hAlign = 'LEFT'
		fotos.append(f)
		fotos.append(Spacer(1, 0.25*inch))"""
		fotos.append(header)#agrega encabezado del doc a lista
		headings = ('Nombre', 'Anio', 'descripcion', 'Procede')#cabecera de tabla
	   
		allclientes = [(Paragraph(p.nombre, style1), p.anio, Paragraph(p.descripcion, style1), Paragraph(str(p.procedencia),styles['Normal'])) for p in Foto.objects.all()]#registros
		t = Table([headings] + allclientes,colWidths=(50*mm, 10*mm, 70*mm, 30*mm),repeatRows=1)#crea tabla
		#t = Table([headings] + allclientes,repeatRows=1)#crea tabla
		t.setStyle(TableStyle(
			[
			('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
			('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
			('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
			]
			))#estilos tabla
		fotos.append(t) #agrega tabla a lista

		#doc.build(fotos) #genera doc en base a lista
		doc.multiBuild(fotos, canvasmaker=FooterCanvas)
		response.write(buff.getvalue()) #imprimimos el doc que sta en el buffer(pdf)
		buff.close()#cerramos buffer
		return response #retornamos pdf

	def get_context_data(self, **kwargs):
		#context = super(Home, self).get_context_data(**kwargs)
		context = super().get_context_data(**kwargs)
		context['objetos'] = Foto.objects.all()#pasamos las fotos
		context['subobjetos'] = Tipofoto.objects.all()#pasamos los tipos
		return context




class Crearfoto(CreateView):
	form_class= FotoForm
	#template_name="fotos/formulariofoto.html"
	template_name="fotos/formulariofoto.html"
	model = Foto
	#fields = '__all__'

class Actualizafoto(UpdateView):
	form_class= FotoForm
	#template_name="fotos/formularioactualizafoto.html"
	template_name="fotos/formulariofoto.html"
	model = Foto
	#fields = '__all__'
class Eliminafoto(DeleteView):
	form_class= FotoForm
	template_name="fotos/formularioeliminafoto1.html"
	model = Foto
	success_url = reverse_lazy("listar")
	def delete(self, request, *args, **kwargs):
		self.object = self.get_object()
		if os.path.exists(str(self.object.ruta)) and os.path.isfile(str(self.object.ruta)):
			os.remove(str(self.object.ruta))
		success_url = self.get_success_url()
		self.object.delete()
		return HttpResponseRedirect(success_url)





class Reportefoto(ListView):
	template_name="reporte_foto.html"
	model = Foto

def listartipos(request):
	#filtro=request.GET['tipo']
	data = Tipofoto.objects.all()
	libros=[tiposerializado(libro) for libro in data]
	return HttpResponse(json.dumps(libros), content_type='application/json')
	


def listarfiltros(request):
	#filtro=request.GET['tipo']
	filtro=request.POST['tipo']
	#data=Foto.objects.all()
	#data=Foto.objects.filter(id=3)
	#tipofoto = Tipofoto.objects.first()
	tipofoto = Tipofoto.objects.get(nombre=filtro)
	data=tipofoto.foto_set.all()
	#data=Foto.objects.get(nombre__exact='Santa Teresita')
	libros=[libroserializado(libro) for libro in data]
	#data="hi"
	#data1 = serializers.serialize("json", data, fields=(nombre))
	#data =filtro
	#data = {'nombre': 'Eduardo', 'dni':'40703678', 'tel':'976649322' }
	#return HttpResponse(json.dumps(data))
	return HttpResponse(json.dumps(libros), content_type='application/json')
	#return HttpResponse(data1, content_type='application/json')
	#return HttpResponse(json.dumps({'valido': False}), content_type='application/javascript')
	#return JsonResponse(json.dumps(fotos), safe=False)

def libroserializado(libro):
	return {'nombre': libro.nombre, 'anio': libro.anio, 'id': libro.id, 'descripcion': libro.descripcion}

def tiposerializado(libro):
	return {'nombre': libro.nombre}



class FotoportipoPdf(Pdf):
	template_name="fotos_tipo.html"
	model = Foto
	modelcampo="tipofoto"
	maestro = Tipofoto
	maestrocampo="nombre" 
	lista_cabeceras=['Nombre', 'Año', 'Descripción', 'Procede']
	lista_campos=['nombre', 'anio', 'descripcion', 'procedencia']
	
	#lista_datos=["valor1","valor2","valor3","valor4"]
	anchocol=(50*mm, 15*mm, 70*mm, 30*mm)# normal:165mm max:195
	titulo="LISTADO DE FOTOS POR "

	

class Pdf(TemplateView):
	template_name="fotos_tipo.html"
	model = Foto
	modelcampo="tipofoto"
	maestro = Tipofoto
	maestrocampo="nombre" 
	lista_cabeceras=['Nombre5', 'Año', 'Descripción', 'Procede']
	lista_campos=['nombre', 'anio', 'descripcion', 'procedencia']
	lista_datos=[[],[],[],[]]
	#lista_datos=["valor1","valor2","valor3","valor4"]
	anchocol=(50*mm, 15*mm, 70*mm, 30*mm)
	titulo="LISTADO DE FOTOS POR "

	def get(self, request, *args, **kwargs):
		response = HttpResponse(content_type='application/pdf')#tipo de rpta
		#pdf_name = "fotos1.pdf"  # nombre para la descarga
		#response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name #descarga con el nombre indicado
		buff = BytesIO() #almacena el doc
		#c = canvas.Canvas(buff)
		doc = SimpleDocTemplate(buff,
								pagesize=A4,
								rightMargin=50,#miderecha
								leftMargin=50,#mi izquierda
								topMargin=60,
								bottomMargin=70,#inferior hoja                              
								)#plantilla del doc enlazado al buffer
		fotos = []#lista para generar doc
		styles = getSampleStyleSheet()#estilos
		style = styles['Title']
		style1= styles['Normal']
		style.spaceAfter = 0.3 * inch

		style.fontSize = 21
		style.fontName = "Times-Roman"
		style.textColor = "Blue"

		#style.alignment=TA_JUSTIFY
		"""p = ParagraphStyle('parrafos', 
						   
						   fontSize = 28,
						   fontName="Times-Roman")"""
		

		filtro = self.kwargs.get('filtro') # El mismo nombre que en tu URL<nombre>
		#pk = self.kwargs.get('pk') # El mismo nombre que en tu URL

		fotos.append(Spacer(1, 0.25*inch))#spacio antes del titulo
		header = Paragraph(self.titulo+filtro, style)#encabezado doc
		#header = Paragraph("Listado de Fotos", styles['Heading1'])#encabezado doc
		#header = Paragraph("Listado de Fotos", p)#encabezado doc
		
		"""f=Image('static/img/admin.png', width=30, height=30)
		f.hAlign = 'LEFT'
		fotos.append(f)
		fotos.append(Spacer(1, 0.25*inch))"""
		fotos.append(header)#agrega encabezado del doc a lista
		headings = self.lista_cabeceras#cabecera de tabla
		#headings = (self.campo1, self.campo2, self.campo3, self.campo4)#cabecera de tabla
		#filtro=request.GET['tipo']
		
		kwargs1 = {}
		kwargs1[self.maestrocampo] = filtro
	
		#eventos = Evento.objects.filter(**kargs) w
		objetofk = self.maestro.objects.get(**kwargs1)
		#tipofoto = self.maestro.objects.get(nombre=filtro)
		kwargs2 = {}
		kwargs2[self.modelcampo] = objetofk
		data=self.model.objects.filter(**kwargs2)
		#data=self.model.objects.filter(tipofoto=tipofoto)
		#data=tipofoto.foto_set.all()
		allclientes=[]
		
		for p in data:
			lista1=self.lista_datos
			
				#if f.name==self.lista_campos[0]:
			for campos in range(len(self.lista_campos)):

				#campo1=[Paragraph(str(p.campo(self.lista_campos[0])), style1)]
				lista1[campos]=[Paragraph(str(p.campo(self.lista_campos[campos])), style1)]
			
			registro=[lista1[0],lista1[1],lista1[2],lista1[3]]
			allclientes.append(registro)
			#allclientes.append([lista1[0],lista1[1],lista1[2],lista1[3]])#agrega cada registro en una lista extendiendola
			#allclientes.extend([lista])
			#allclientes.extend([(Paragraph(p.nombre, style1), p.anio, Paragraph(p.descripcion, style1), Paragraph(str(p.procedencia),styles['Normal']))])

		#allclientes = [(Paragraph(p.nombre, style1), p.anio, Paragraph(p.descripcion, style1), Paragraph(str(p.procedencia),styles['Normal'])) for p in data]#registros
		t = Table([headings] + allclientes,colWidths=self.anchocol,repeatRows=1)#crea tabla
		#t = Table([headings] + allclientes,repeatRows=1)#crea tabla
		t.setStyle(TableStyle(
			[
			('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
			('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
			('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
			]
			))#estilos tabla
		fotos.append(t) #agrega tabla a lista

		#doc.build(fotos) #genera doc en base a lista
		doc.title = self.titulo+filtro # si no se coloca aparece anonymous
		doc.multiBuild(fotos, canvasmaker=FooterCanvas)

		response.write(buff.getvalue()) #imprimimos el doc que sta en el buffer(pdf)
		buff.close()#cerramos buffer
		return response #retornamos pdf
		#return render(response, self.template_name)




class Reportefototipo(ListView):
	template_name="fotos_tipo.html"
	model = Foto
	def pdf_Platypus(request,filtro):
		response = HttpResponse(content_type='application/pdf')#tipo de rpta
		#pdf_name = "fotos1.pdf"  # nombre para la descarga
		#response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name #descarga con el nombre indicado
		buff = BytesIO() #almacena el doc
		#c = canvas.Canvas(buff)
		doc = SimpleDocTemplate(buff,
								pagesize=A4,
								rightMargin=50,#miderecha
								leftMargin=50,#mi izquierda
								topMargin=60,
								bottomMargin=70,#inferior hoja                              
								)#plantilla del doc enlazado al buffer
		fotos = []#lista para generar doc
		styles = getSampleStyleSheet()#estilos
		style = styles['Title']
		style1= styles['Normal']
		style.spaceAfter = 0.3 * inch
		style.fontSize = 21
		style.fontName = "Times-Roman"
		style.textColor = "Blue"

		#style.alignment=TA_JUSTIFY
		"""p = ParagraphStyle('parrafos', 
						   
						   fontSize = 28,
						   fontName="Times-Roman")"""
		header = Paragraph("Listado de Fotos por "+filtro, style)#encabezado doc
		#header = Paragraph("Listado de Fotos", styles['Heading1'])#encabezado doc
		#header = Paragraph("Listado de Fotos", p)#encabezado doc
		
		"""f=Image('static/img/admin.png', width=30, height=30)
		f.hAlign = 'LEFT'
		fotos.append(f)
		fotos.append(Spacer(1, 0.25*inch))"""
		fotos.append(header)#agrega encabezado del doc a lista
		headings = ('Nombre', 'Anio', 'descripcion', 'Procede')#cabecera de tabla
		#filtro=request.GET['tipo']
		
		tipofoto = Tipofoto.objects.get(nombre=filtro)
		data=tipofoto.foto_set.all()
	   
		allclientes = [(Paragraph(p.nombre, style1), p.anio, Paragraph(p.descripcion, style1), Paragraph(str(p.procedencia),styles['Normal'])) for p in data]#registros
		t = Table([headings] + allclientes,colWidths=(50*mm, 10*mm, 70*mm, 30*mm),repeatRows=1)#crea tabla
		#t = Table([headings] + allclientes,repeatRows=1)#crea tabla
		t.setStyle(TableStyle(
			[
			('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
			('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
			('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
			]
			))#estilos tabla
		fotos.append(t) #agrega tabla a lista

		#doc.build(fotos) #genera doc en base a lista
		doc.title = "Listado de Fotos por "+filtro # si no se coloca aparece anonymous
		doc.multiBuild(fotos, canvasmaker=FooterCanvas)

		response.write(buff.getvalue()) #imprimimos el doc que sta en el buffer(pdf)
		buff.close()#cerramos buffer
		return response #retornamos pdf

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteFotosExcel(ListView):
	template_name="fotos_tipo.html"
	model = Foto
	#Usamos el método get para generar el archivo excel
	def excel_openpyxl(request,filtro):
	#def get(self, request, *args, **kwargs):
		#Obtenemos todas las personas de nuestra base de datos
		tipofoto = Tipofoto.objects.get(nombre=filtro)
		fotos=tipofoto.foto_set.all()

		#fotos = Foto.objects.all()
		#Creamos el libro de trabajo
		wb = Workbook()
		#Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
		ws = wb.active
		#En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
		ws['B1'] = 'Listado de Fotos por '+filtro
		#Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
		ws.merge_cells('B1:E1')
		#Creamos los encabezados desde la celda B3 hasta la E3
		ws['B3'] = 'NOMBRE'
		ws['C3'] = 'AÑO'
		ws['D3'] = 'DESCRIPCION'
		ws['E3'] = 'PROCEDENCIA'       
		cont=4
		#Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
		for foto in fotos:
			ws.cell(row=cont,column=2).value = foto.nombre
			ws.cell(row=cont,column=3).value = foto.anio
			ws.cell(row=cont,column=4).value = foto.descripcion
			ws.cell(row=cont,column=5).value = str(foto.procedencia)
			cont = cont + 1
		#Establecemos el nombre del archivo
		nombre_archivo ="ReporteFotosExcel.xlsx"
		#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
		response = HttpResponse(content_type="application/ms-excel") 
		contenido = "attachment; filename={0}".format(nombre_archivo)
		response["Content-Disposition"] = contenido
		wb.save(response)
		return response
class Eliminatipofoto(DeleteView):
	form_class= TipoFotoForm
	template_name="fotos/formularioeliminatipofoto.html"
	model = Tipofoto
	success_url = reverse_lazy("lista_tipo_foto")

class Eliminaprocedencia(DeleteView):
	form_class= ProcedenciaForm
	template_name="fotos/formularioeliminaprocedencia.html"
	model = Procedencia
	success_url = reverse_lazy("lista_procedencia")
