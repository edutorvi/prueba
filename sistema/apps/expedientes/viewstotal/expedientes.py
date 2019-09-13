# -*- coding: utf-8 -*-
from io import BytesIO#PARA REPORTLAB

from django.views.generic import TemplateView, CreateView, UpdateView, DeleteView, ListView
from apps.expedientes.models import Expediente, Tipoexpediente, Serie
from apps.expedientes.forms import ExpedienteForm,TipoExpedienteForm,SerieForm
from django.views import  View
from django.urls import reverse_lazy
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.http import JsonResponse
from django.core import serializers

from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
from reportlab.lib.styles import getSampleStyleSheet#estilosparrafo
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
#from reportlab.lib.pagesizes import LETTER
from reportlab.platypus import Table
from reportlab.pdfgen import canvas
#Librerias reportlab a usar:
from reportlab.platypus import (BaseDocTemplate, Frame,  
					NextPageTemplate, PageBreak, PageTemplate)
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4

from reportlab.lib.styles import ParagraphStyle

from reportlab.lib.units import mm
from reportlab.platypus import Image
from reportlab.platypus.flowables import Spacer


from django.shortcuts import render
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
import os
from django.http import HttpResponseRedirect

from django.db import IntegrityError


class FooterCanvas(canvas.Canvas):

	def __init__(self, *args, **kwargs):
		canvas.Canvas.__init__(self, *args, **kwargs)
		self.pages = []

	def showPage(self):
		self.pages.append(dict(self.__dict__))
		self._startPage()

	def save(self):
		page_count = len(self.pages)
		for page in self.pages:
			self.__dict__.update(page)
			self.draw_canvas(page_count)
			canvas.Canvas.showPage(self)
		canvas.Canvas.save(self)

	def draw_canvas(self, page_count):
		page1 = "Pág %s de %s" % (self._pageNumber, page_count)
		page2 = "Gobierno Regional de Cajamarca"
		x1 = 128
		x2 = 500
		x2 = 461
		self.saveState()
		self.setStrokeColorRGB(0, 0, 0)
		self.setLineWidth(0.5)
		self.line(66, 71, A4[0] - 66, 71)#linea footer
		self.setFont('Times-Roman', 10)
		self.drawString(A4[0]-x1, 59, page1)#texto footer
		
		self.setFont('Times-Roman', 12)
		self.drawString(A4[0]-x2, 785, page2)#texto cabecera
		self.line(66, 780, A4[0] - 66, 780)#linea cabecera

		#self = canvas.Canvas(filename, bottomup=0)
		f=Image('static/img/logo_Region_Cajamarca2019.png', width=147, height=50)#imagen
		#f=Image('static/img/gr2019.png', width=73, height=25)#imagen
		f.drawOn(self, 75, 780)#posicion DE LA IMAGEN en el encabezado


		self.restoreState()
	
class Listarexpediente(TemplateView):
	template_name = "listarexpediente.html"
	def pdf_Platypus(request):
		response = HttpResponse(content_type='application/pdf')#tipo de rpta
		#pdf_name = "expedientes1.pdf"  # nombre para la descarga
		#response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name #descarga con el nombre indicado
		buff = BytesIO() #almacena el doc
		#c = canvas.Canvas(buff)
		doc = SimpleDocTemplate(buff,
								pagesize=A4,
								rightMargin=50,#miderecha
								leftMargin=50,#mi izquierda
								topMargin=60,
								bottomMargin=70,#inferior hoja	                            
								)#plantilla del doc enlazado al buffer
		expedientes = []#lista para generar doc
		styles = getSampleStyleSheet()#estilos
		style = styles['Title']
		style1= styles['Normal']
		style.spaceAfter = 0.3 * inch
		style.fontSize = 21
		style.fontName = "Times-Roman"
		style.textColor = "Blue"

		#style.alignment=TA_JUSTIFY
		"""p = ParagraphStyle('parrafos', 
						   
						   fontSize = 28,
						   fontName="Times-Roman")"""
		header = Paragraph("Listado de expedientes", style)#encabezado doc
		#header = Paragraph("Listado de expedientes", styles['Heading1'])#encabezado doc
		#header = Paragraph("Listado de expedientes", p)#encabezado doc
		
		"""f=Image('static/img/admin.png', width=30, height=30)
		f.hAlign = 'LEFT'
		expedientes.append(f)
		expedientes.append(Spacer(1, 0.25*inch))"""
		expedientes.append(header)#agrega encabezado del doc a lista
		headings = ('Nombre', 'Anio', 'descripcion', 'Procede')#cabecera de tabla
	   
		allclientes = [(Paragraph(p.nombre, style1), p.anio, Paragraph(p.descripcion, style1), Paragraph(str(p.serie),styles['Normal'])) for p in expediente.objects.all()]#registros
		t = Table([headings] + allclientes,colWidths=(50*mm, 10*mm, 70*mm, 30*mm),repeatRows=1)#crea tabla
		#t = Table([headings] + allclientes,repeatRows=1)#crea tabla
		t.setStyle(TableStyle(
			[
			('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
			('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
			('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
			]
			))#estilos tabla
		expedientes.append(t) #agrega tabla a lista

		#doc.build(expedientes) #genera doc en base a lista
		doc.multiBuild(expedientes, canvasmaker=FooterCanvas)
		response.write(buff.getvalue()) #imprimimos el doc que sta en el buffer(pdf)
		buff.close()#cerramos buffer
		return response #retornamos pdf

	def get_context_data(self, **kwargs):
		#context = super(Home, self).get_context_data(**kwargs)
		context = super().get_context_data(**kwargs)
		context['objetos'] = Expediente.objects.all()#pasamos las expedientes
		#context['subobjetos'] = Tipoexpediente.objects.all()#pasamos los tipos
		return context




class Crearexpediente(CreateView):
	form_class= ExpedienteForm
	template_name="expedientes/formularioexpediente.html"
	model = Expediente

	#fields = '__all__'

class Actualizaexpediente(UpdateView):
	form_class= ExpedienteForm
	template_name="expedientes/formularioexpediente.html"
	model = Expediente
	#fields = '__all__'
class Eliminaexpediente(DeleteView):
	form_class= ExpedienteForm
	template_name="expedientes/formularioeliminaexpediente.html"
	model = Expediente
	success_url = reverse_lazy("lista_expediente")
	def delete(self, request, *args, **kwargs):
		self.object = self.get_object()
		if os.path.exists(str(self.object.ruta)) and os.path.isfile(str(self.object.ruta)):
			os.remove(str(self.object.ruta))
		success_url = self.get_success_url()
		self.object.delete()
		return HttpResponseRedirect(success_url)





class Reporteexpediente(ListView):
	template_name="reporte_expediente.html"
	model = Expediente

def listartipos(request):
	#filtro=request.GET['tipo']
	data = Tipoexpediente.objects.all()
	libros=[tiposerializado(libro) for libro in data]
	return HttpResponse(json.dumps(libros), content_type='application/json')
	


def listarfiltros(request):
	filtro=request.GET['tipo']
	#data=expediente.objects.all()
	#data=expediente.objects.filter(id=3)
	#tipoexpediente = Tipoexpediente.objects.first()
	tipoexpediente = Tipoexpediente.objects.get(nombre=filtro)
	data=tipoexpediente.expediente_set.all()
	#data=expediente.objects.get(nombre__exact='Santa Teresita')
	libros=[libroserializado(libro) for libro in data]
	#data="hi"
	#data1 = serializers.serialize("json", data, fields=(nombre))
	#data =filtro
	#data = {'nombre': 'Eduardo', 'dni':'40703678', 'tel':'976649322' }
	#return HttpResponse(json.dumps(data))
	return HttpResponse(json.dumps(libros), content_type='application/json')
	#return HttpResponse(data1, content_type='application/json')
	#return HttpResponse(json.dumps({'valido': False}), content_type='application/javascript')
	#return JsonResponse(json.dumps(expedientes), safe=False)

def libroserializado(libro):
	return {'nombre': libro.nombre, 'anio': libro.anio, 'id': libro.id, 'descripcion': libro.descripcion}

def tiposerializado(libro):
	return {'nombre': libro.nombre}


class Reporteexpedientetipo(ListView):
	template_name="expediente_tipo.html"
	model = Expediente
	def pdf_Platypus(request,filtro):
		response = HttpResponse(content_type='application/pdf')#tipo de rpta
		#pdf_name = "expedientes1.pdf"  # nombre para la descarga
		#response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name #descarga con el nombre indicado
		buff = BytesIO() #almacena el doc
		#c = canvas.Canvas(buff)
		doc = SimpleDocTemplate(buff,
								pagesize=A4,
								rightMargin=50,#miderecha
								leftMargin=50,#mi izquierda
								topMargin=60,
								bottomMargin=70,#inferior hoja                              
								)#plantilla del doc enlazado al buffer
		expedientes = []#lista para generar doc
		styles = getSampleStyleSheet()#estilos
		style = styles['Title']
		style1= styles['Normal']
		style.spaceAfter = 0.3 * inch
		style.fontSize = 21
		style.fontName = "Times-Roman"
		style.textColor = "Blue"

		#style.alignment=TA_JUSTIFY
		"""p = ParagraphStyle('parrafos', 
						   
						   fontSize = 28,
						   fontName="Times-Roman")"""
		header = Paragraph("Listado de expedientes por "+filtro, style)#encabezado doc
		#header = Paragraph("Listado de expedientes", styles['Heading1'])#encabezado doc
		#header = Paragraph("Listado de expedientes", p)#encabezado doc
		
		"""f=Image('static/img/admin.png', width=30, height=30)
		f.hAlign = 'LEFT'
		expedientes.append(f)
		expedientes.append(Spacer(1, 0.25*inch))"""
		expedientes.append(header)#agrega encabezado del doc a lista
		headings = ('Nombre', 'Anio', 'descripcion', 'Procede')#cabecera de tabla
		#filtro=request.GET['tipo']
		
		tipoexpediente = Tipoexpediente.objects.get(nombre=filtro)
		data=tipoexpediente.expediente_set.all()
	   
		allclientes = [(Paragraph(p.nombre, style1), p.anio, Paragraph(p.descripcion, style1), Paragraph(str(p.serie),styles['Normal'])) for p in data]#registros
		t = Table([headings] + allclientes,colWidths=(50*mm, 10*mm, 70*mm, 30*mm),repeatRows=1)#crea tabla
		#t = Table([headings] + allclientes,repeatRows=1)#crea tabla
		t.setStyle(TableStyle(
			[
			('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
			('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
			('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
			]
			))#estilos tabla
		expedientes.append(t) #agrega tabla a lista

		#doc.build(expedientes) #genera doc en base a lista
		doc.title = "Listado de expedientes por "+filtro # si no se coloca aparece anonymous
		doc.multiBuild(expedientes, canvasmaker=FooterCanvas)

		response.write(buff.getvalue()) #imprimimos el doc que sta en el buffer(pdf)
		buff.close()#cerramos buffer
		return response #retornamos pdf

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteexpedientesExcel(ListView):
	template_name="expediente_tipo.html"
	model = Expediente
	#Usamos el método get para generar el archivo excel
	def excel_openpyxl(request,filtro):
	#def get(self, request, *args, **kwargs):
		#Obtenemos todas las personas de nuestra base de datos
		tipoexpediente = Tipoexpediente.objects.get(nombre=filtro)
		expedientes=tipoexpediente.expediente_set.all()

		#expedientes = expediente.objects.all()
		#Creamos el libro de trabajo
		wb = Workbook()
		#Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
		ws = wb.active
		#En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
		ws['B1'] = 'Listado de expedientes por '+filtro
		#Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
		ws.merge_cells('B1:E1')
		#Creamos los encabezados desde la celda B3 hasta la E3
		ws['B3'] = 'NOMBRE'
		ws['C3'] = 'AÑO'
		ws['D3'] = 'DESCRIPCION'
		ws['E3'] = 'serie'       
		cont=4
		#Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
		for expediente in expedientes:
			ws.cell(row=cont,column=2).value = expediente.nombre
			ws.cell(row=cont,column=3).value = expediente.anio
			ws.cell(row=cont,column=4).value = expediente.descripcion
			ws.cell(row=cont,column=5).value = str(expediente.serie)
			cont = cont + 1
		#Establecemos el nombre del archivo
		nombre_archivo ="ReporteexpedientesExcel.xlsx"
		#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
		response = HttpResponse(content_type="application/ms-excel") 
		contenido = "attachment; filename={0}".format(nombre_archivo)
		response["Content-Disposition"] = contenido
		wb.save(response)
		return response
class Eliminatipoexpediente(DeleteView):
	form_class= TipoExpedienteForm
	template_name="expedientes/formularioeliminatipoexpediente.html"
	model = Tipoexpediente
	success_url = reverse_lazy("lista_tipo_expediente")

class Crearserie(CreateView):
	form_class= SerieForm
	template_name="expedientes/formularioserie.html"
	model = Serie
	def form_invalid(self, form):
		response = super().form_invalid(form)
		if self.request.is_ajax():
			#return JsonResponse(form.errors, status=400)
			return HttpResponse(form.errors.as_json(), status = 400, content_type='application/json')
			#return JsonResponse(form.errors)
		else:
			return response
	def form_valid(self, form):
		# We make sure to call the parent's form_valid() method because
		# it might do some processing (in the case of CreateView, it will
		# call form.save() for example).
		response = super().form_valid(form)
		if self.request.is_ajax():
			data = {
				'pk': self.object.pk,
			}
			return JsonResponse(data)
		else:
			return response


	#def form_invalid(self,form):
	#	return super(Crearserie, self).form_invalid(form)
	# def form_valid(self, form):
	#     return super(Crearserie, self).form_valid(form)
		
	# try:
	# 	return super(Crearserie, self).form_valid(form)
	# except IntegrityError:
	# 	form.add_error(,"errordupli")
	# 	#raise forms.ValidationError("Ingrese 4 caracteresperu")
	# 	#return HttpResponse("ERROR: Kumquat already exists!")
	# 	return super(Crearserie, self).form_invalid(form)
	

class Eliminaserie(DeleteView):
	form_class= SerieForm
	template_name="expedientes/formularioeliminaserie.html"
	model = Serie
	success_url = reverse_lazy("lista_serie")
